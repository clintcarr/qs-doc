from openpyxl import Workbook
from openpyxl.compat import range
# from tqdm import tqdm
from openpyxl.utils import get_column_letter
import get_qlik_sense
# import server_details
import argparse
import datetime

year = datetime.date.today().year
month = datetime.date.today().month
day = datetime.date.today().day



parser = argparse.ArgumentParser()
parser.add_argument('--server', help='Qlik Sense Server to connect to.')
parser.add_argument('--certs', help='Path to certificates.')
parser.add_argument('--user', help='Username in format domain\\user')
parser.add_argument('--wmi', help='True/False, set to True to collect WMI information from servers. This switch requires windows authentication (username and password)')
parser.add_argument('--password', help='Password of user.')
args = parser.parse_args()

wb = Workbook()
dest_filename = 'empty_book.xlsx'

def nodes():
    ws1 = wb.active
    ws1.title = 'Server Nodes'
    servernodes = get_qlik_sense.get_servernode()
    node_cols = ['node name', 'hostname', 'service cluster']
    num_of_nodes = len(servernodes)
    r = 1
    for item in node_cols:
        ws1.cell(row=1, column=r).value = item
        r += 1

    for row in range(num_of_nodes):
        ws1.append(servernodes[row])


def node_config():
    """
    Collect and collate the node configuration in the Qlik Sense Site
    """
    ws2 = wb.create_sheet(title='Node Configuration')
    node_metrics = ['host name','central','purpose', 'engine', 'proxy', 'printing', 'scheduler','name']
    nodes = get_qlik_sense.get_nodeconfig()
    num_of_nodes = len(nodes)
    num_of_node_metrics = len(node_metrics)
    r = 1
    for item in node_metrics:
        ws2.cell(row=1, column=r).value = item
        r += 1
    for row in range(num_of_nodes):
        ws2.append(nodes[row])

def default_rules():
    """
    Collect and collate the system rules in the Qlik Sense Site
    """
    ws3 = wb.create_sheet(title='Default System Rules')
    rule_cols = ['name', 'rule', 'resource filter', 'actions', 'disabled']
   
    systemrules = get_qlik_sense.get_systemrules('Default')
    num_of_systemrules = len(systemrules)
    r = 1
    for item in rule_cols:
        ws3.cell(row=1, column=r).value = item
        r += 1
    for row in range(num_of_systemrules):
        ws3.append(systemrules[row])

def custom_rules():
    """
    Collect and collate the system rules in the Qlik Sense Site
    """
    ws4 = wb.create_sheet(title='Custom System Rules')
    rule_cols = ['name', 'rule', 'resource filter', 'actions', 'disabled']
    systemrules = get_qlik_sense.get_systemrules('Custom')
    num_of_systemrules = len(systemrules)
    r = 1
    for item in rule_cols:
        ws4.cell(row=1, column=r).value = item
        r += 1
    
    for row in range(num_of_systemrules):
        ws4.append(systemrules[row])

def readonly_rules():
    """
    Collect and collate the system rules in the Qlik Sense Site
    """
    rule_cols = ['name', 'rule', 'resource filter', 'actions', 'disabled']
    ws5 = wb.create_sheet(title='Read Only Rules')
   
    systemrules = get_qlik_sense.get_systemrules('ReadOnly')
    num_of_systemrules = len(systemrules)
    r = 1
    for item in rule_cols:
        ws5.cell(row=1, column=r).value = item
        r += 1
    for row in range(num_of_systemrules):
        ws5.append(systemrules[row])

def apps():
    """
    Collect and collate the system rules in the Qlik Sense Site
    """
    app_cols = ['name', 'description', 'publish time', 'stream', 'file size', 'owner user ID', 'owner user Name']
    ws6 = wb.create_sheet(title='Applications')
    apps = get_qlik_sense.get_apps()
    num_of_apps = len(apps)
    r = 1
    for item in app_cols:
        ws6.cell(row=1, column=r).value = item
        r += 1
    for row in range(num_of_apps):
        ws6.append(apps[row])

def streams():
    ws7 = wb.create_sheet(title='Streams')
    streams = get_qlik_sense.get_streams()
    num_of_streams = len(streams)
    stream_cols = ['name']
    streamname = []
    for row in range(num_of_streams):
        streamname.append (streams[row]['name'])
    r = 1
    for item in stream_cols:
        ws7.cell(row=r, column=1).value = item
        r += 1
    r = 2
    for item in streamname:
        ws7.cell(row=r, column=1).value = item
        r += 1

def dataconnections():
    """
    Collect and collate the data connections
    """
    ws8 = wb.create_sheet(title='Data Connections')
    dc_cols = ['name', 'connection string', 'type']
    connections = get_qlik_sense.get_dataconnections()
    num_of_connections = len(connections)
    r = 1
    for item in dc_cols:
        ws8.cell(row=1, column=r).value = item
        r += 1
    for row in range(num_of_connections):
        ws8.append(connections[row])

def engine():
    """
    Collect and collate the engines within the site
    """
    ws9 = wb.create_sheet(title='Engines')
    engine_cols = ['customProperties','listenerPorts','autosaveInterval','documentTimeout','documentDirectory',
                      'workingSetSizeLoPct', 'workingSetSizeHiPct', 'cpuThrottlePercentage', 'maxCoreMaskPersisted', 'maxCoreMask',
                      'maxCoreMaskHiPersisted', 'maxCoreMaskHi','objectTimeLimitSec', 'exportTimeLimitSec', 'reloadTimeLimitSec',
                      'hyperCubeMemoryLimit', 'exportMemoryLimit', 'reloadMemoryLimit','hostname', 'name']
    enginenodes = get_qlik_sense.get_engine()
    num_of_engines = len(enginenodes)
    r = 1
    for item in engine_cols:
        ws9.cell(row=1, column=r).value = item
        r += 1
    for item in range (num_of_engines):
        r = 1
        for metric in range(len(enginenodes[item])):
            ws9.cell(row=item+2, column = r).value = str((enginenodes[item][metric]))
            r += 1

def scheduler():
    """
    Collect and collate the schedulers within the site
    """
    ws10 = wb.create_sheet(title='Schedulers')
    scheduler_cols = ['customProperties', 
                      'schedulerServiceType', 
                      'maxConcurrentEngines', 
                      'engineTimeout', 
                      'tags', 
                      'hostname',
                      'name']
    schedulernodes = get_qlik_sense.get_scheduler()
    num_of_schedulers = len(schedulernodes)
    r = 1
    for item in scheduler_cols:
        ws10.cell(row=1, column=r).value = item
        r += 1
    for item in range (num_of_schedulers):
        r = 1
        for metric in range(len(schedulernodes[item])):
            ws10.cell(row=item+2, column = r).value = str((schedulernodes[item][metric]))
            r += 1

def proxy():
    """
    Collect and collate the schedulers within the site
    """
    ws11 = wb.create_sheet(title='Proxy')
    proxy_cols = ['customProperties',
                  'listenPort',
                  'allowHttp',
                  'unencryptedListenPort',
                  'authenticationListenPort',
                  'kerberosAuthentication',
                  'unencryptedAuthenticationListenPort',
                  'keepAliveTimeoutSeconds',
                  'maxHeaderSizeBytes',
                  'maxHeaderLines',
                  'hostName',
                  'name']

    proxynodes = get_qlik_sense.get_proxy()
    num_of_proxy = len(proxynodes)
    r = 1
    for item in proxy_cols:
        ws11.cell(row=1, column=r).value = item
        r += 1
    for item in range (num_of_proxy):
        r = 1
        for metric in range(len(proxynodes[item])):
            ws11.cell(row=item+2, column = r).value = str((proxynodes[item][metric]))
            r += 1

def virtual_proxy():
    ws12 = wb.create_sheet(title='Virtual Proxy')
    virtualproxy_cols= ['description',
                        'prefix',
                        'authenticationModuleRedirectUri',
                        'sessionModuleBaseUri',
                        'loadBalancingModuleBaseUri',
                        'authenticationMethod',
                        'headerAuthenticationMode',
                        'headerAuthenticationHeaderName',
                        'headerAuthenticationStaticUserDirectory',
                        'headerAuthenticationDynamicUserDirectory',
                        'anonymousAccessMode',
                        'windowsAuthenticationEnabledDevicePattern',
                        'sessionCookieHeaderName',
                        'sessionCookieDomain',
                        'additionalResponseHeaders',
                        'sessionInactivityTimeout',
                        'extendedSecurityEnvironment',
                        'websocketCrossOriginWhiteList',
                        'defaultVirtualProxy',
                        'tags',
                        'samlMetadataIdP',
                        'samlHostUri',
                        'samlEntityId',
                        'samlAttributeUserId',
                        'samlAttributeUserDirectory',
                        'samlAttributeSigningAlgorithm',
                        'samlAttributeMap',
                        'magicLinkHostUri',
                        'magicLinkFriendlyName',
                        'name']
    virtualproxynodes = get_qlik_sense.get_virtualproxy()
    num_of_virtualproxys = len(virtualproxynodes)
    r = 1
    for item in virtualproxy_cols:
        ws12.cell(row=1, column=r).value = item
        r += 1
    for item in range(num_of_virtualproxys):
        r = 1
        for metric in range(len(virtualproxynodes[item])):
            ws12.cell(row=item+2, column=r).value = str((virtualproxynodes[item][metric]))
            r += 1

def virtual_proxylb():

    ws13 = wb.create_sheet(title='Virtual Proxy Load Balancers')
    virtualproxylb_cols = ['node', 'load balance nodes']
    virtualproxyslb = get_qlik_sense.get_vploadbalancers()
    num_of_virtualproxyslb = len(virtualproxyslb)
    r = 1
    for item in virtualproxylb_cols:
        ws13.cell(row=1, column=r).value = item
        r += 1

    for item in range(num_of_virtualproxyslb):
        ws13.cell(row=item+2, column=1).value = str((virtualproxyslb[item]['type']))
        ws13.cell(row=item+2, column=2).value = str((virtualproxyslb[item]['items']))

def custom_properties():
    ws14 = wb.create_sheet(title='Custom Properties')
    customprop_cols = ['name', 'choice values', 'object types']
    customproperties = get_qlik_sense.get_customprop()
    num_of_customproperties = len(customproperties)
    r = 1
    for item in customprop_cols:
        ws14.cell(row=1, column=r).value = item
        r += 1
    for item in range(num_of_customproperties):
        ws14.cell(row=item+2, column=1).value = str(customproperties[item][0])
        ws14.cell(row=item+2, column=2).value = str(customproperties[item][1])
        ws14.cell(row=item+2, column=3).value = str(customproperties[item][2])

def tags():
    ws15 = wb.create_sheet(title='Tags')
    tags = get_qlik_sense.get_tag()
    num_of_tags = len(tags)
    tag_cols = ['name']
    r = 1
    for item in tag_cols:
        ws15.cell(row=r, column=1).value = item
        r += 1
    r = 2
    for item in tags:
        ws15.cell(row=r, column=1).value = item
        r += 1

def extensions():
    ws16 = wb.create_sheet(title='Extensions')
    extensions = get_qlik_sense.get_extensions()
    num_of_extensions = len(extensions)
    extension_cols = ['name']
    r = 1
    for item in extension_cols:
        ws16.cell(row=r, column=1).value = item
        r += 1
    r = 2
    for item in extensions:
        ws16.cell(row=r, column=1).value = item
        r += 1

def service_cluster():
    ws17 = wb.create_sheet(title='Service Cluster')
    sc_cols = ['name', 
               'root folder',
               'app folder',
               'static content',
               '32bit Connector',
               '64bit Connector','archived logs',
               'database host',
               'database port']
    qs_sc = get_qlik_sense.get_servicecluster()
    num_of_metric = len(qs_sc)
    r = 1
    for item in sc_cols:
        ws17.cell(row=1, column=r).value = item
        r += 1
    r = 1
    for item in qs_sc:
        ws17.cell(row=2, column= r).value = item
        r += 1

def qs_license():
    ws18 = wb.create_sheet(title='License')
    lic_cols = ['lef',
                'serial',
                'name',
                'organization',
                'product',
                'numberOfCores',
                'isExpired',
                'expiredReason',
                'isBlacklisted',
                'isInvalid']
    qs_lic = get_qlik_sense.get_license()
    num_of_metric = len(qs_lic)
    r = 1
    for item in lic_cols:
        ws18.cell(row=1, column=r).value = item
        r += 1
    r = 1
    for item in qs_lic:
        ws18.cell(row=2, column= r).value = str(item)
        r += 1

def userdirectories():
    ws19 = wb.create_sheet(title='User Directory Connectors')
    udc_cols = ['name',
                'user directory name',
                'configured',
                'operational',
                'type',
                'synchronise only loggined in users'
                ]
    userdirectories = get_qlik_sense.get_userdirectory()
    num_of_udc = len(userdirectories)
    r = 1
    for item in udc_cols:
        ws19.cell(row=1, column=r).value = item
        r += 1
    r = 1
    for udc in range(len(userdirectories)):
        ws19.cell(row=udc+2, column=1).value = str(userdirectories[udc][0])
        ws19.cell(row=udc+2, column=2).value = str(userdirectories[udc][1])
        ws19.cell(row=udc+2, column=3).value = str(userdirectories[udc][2])
        ws19.cell(row=udc+2, column=4).value = str(userdirectories[udc][3])
        ws19.cell(row=udc+2, column=5).value = str(userdirectories[udc][4])
        ws19.cell(row=udc+2, column=6).value = str(userdirectories[udc][5])
        r += 1

def wmi():
    try:
        ws20 = wb.create_sheet(title='Machine information')
        wminodes = server_details.auth.wmi_results()
        wmi_cols = ['operating system',
                    'ram free',
                    'total ram',
                    'disk free',
                    'disk total',
                    'cpu name',
                    'number of sockets',
                    'number of logical cores']
        r = 1
        for item in wmi_cols:
            ws20.cell(row=1, column=r).value = item
            r += 1
        r = 1
        for metric in range(len(wminodes)):
            ws20.cell(row=metric+2, column=1).value = str(wminodes[metric][0])
            ws20.cell(row=metric+2, column=2).value = str(wminodes[metric][1])
            ws20.cell(row=metric+2, column=3).value = str(wminodes[metric][2])
            ws20.cell(row=metric+2, column=4).value = str(wminodes[metric][3])
            ws20.cell(row=metric+2, column=5).value = str(wminodes[metric][4])
            ws20.cell(row=metric+2, column=6).value = str(wminodes[metric][5])
            ws20.cell(row=metric+2, column=7).value = str(wminodes[metric][6])
            ws20.cell(row=metric+2, column=8).value = str(wminodes[metric][7])
            r += 1
    except Exception:
        pass

def main():

    nodes()
    if args.wmi:
        wmi()
    node_config()
    service_cluster()
    qs_license()
    engine()
    scheduler()
    proxy()
    virtual_proxy()
    virtual_proxylb()
    default_rules()
    custom_rules()
    readonly_rules()
    userdirectories()
    apps()
    streams()
    dataconnections()
    custom_properties()
    tags()
    extensions()
    wb.save('QSXls_{0}_{1}_{2}.xlsx'.format(year, month, day))
